from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.db.postgres import get_db
from app.services.reports import build_report_data
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

router = APIRouter(prefix="/reports", tags=["reports"])


def get_default_period():
    now = datetime.now()
    start = now.replace(day=1).strftime("%Y-%m-%d")
    end = now.strftime("%Y-%m-%d")
    return start, end


def resolve_period(start: str = None, end: str = None):
    """Если start не задан — берём начало текущего месяца.
    Если end не задан — берём сегодня (даже если start задан)."""
    default_start, _ = get_default_period()
    resolved_start = (start or default_start)[:10]
    resolved_end = (end or datetime.now().strftime("%Y-%m-%d"))[:10]
    return resolved_start, resolved_end


def parse_machine_ids(machine_id: str = None):
    """machine_id может быть пустым (все станции), '35863' (одна) или '35863,35864' (несколько)"""
    if not machine_id:
        return None
    return [int(x.strip()) for x in machine_id.split(",") if x.strip()]


@router.get("/preview", summary="Предпросмотр отчёта для фронта (с пагинацией)")
async def preview_report(
    machine_id: str = None,
    start: str = None,
    end: str = None,
    page: int = 1,
    page_size: int = 10,
    db: AsyncSession = Depends(get_db)
):
    start, end = resolve_period(start, end)

    machine_ids = parse_machine_ids(machine_id)
    totals, days = await build_report_data(db, machine_ids, start, end, refresh_today=True)

    total_days = len(days)
    total_pages = (total_days + page_size - 1) // page_size if total_days else 1

    page = max(1, min(page, total_pages))
    offset = (page - 1) * page_size
    paginated_days = days[offset:offset + page_size]

    return {
        "period": {"start": start, "end": end},
        "machine_ids": machine_ids,
        "totals": totals,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_items": total_days,
            "total_pages": total_pages,
        },
        "days": paginated_days
    }


@router.get("/excel", summary="Скачать финансовый отчёт в Excel")
async def export_excel(
    machine_id: str = None,
    start: str = None,
    end: str = None,
    db: AsyncSession = Depends(get_db)
):
    start, end = resolve_period(start, end)

    machine_ids = parse_machine_ids(machine_id)
    totals, days = await build_report_data(db, machine_ids, start, end, refresh_today=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = [
        "Дата", "Общая сумма", "Наличные", "Безналичные", "Оплата по QR",
        "Банкнотами", "Монетами", "Поступление с ЦПТ",
        "Мобильное приложение (всего)", "Мобильное приложение (бонусы)"
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.append([
        "Всего", totals["total"], totals["cash"], totals["cashless"], totals["qr"],
        totals["banknotes"], totals["coins"], totals["cpt"], totals["mobile_app"], totals["mobile_bonus"]
    ])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for d in days:
        date_short = d["date"][5:]
        ws.append([
            date_short, d["total"], d["cash"], d["cashless"], d["qr"],
            d["banknotes"], d["coins"], d["cpt"], d["mobile_app"], d["mobile_bonus"]
        ])

    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    machine_label = machine_id if machine_id else "all"
    filename = f"report_{machine_label}_{start}_{end}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/programs-report/excel", summary="Скачать отчёт по тарифам в Excel")
async def export_programs_report_excel(
    machine_id: str = None,
    start: str = None,
    end: str = None,
    db: AsyncSession = Depends(get_db)
):
    start, end = resolve_period(start, end)
    machine_ids = parse_machine_ids(machine_id)

    from app.services.reports import build_programs_report_data
    summary, programs = await build_programs_report_data(db, machine_ids, start, end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тарифы"

    ws.append([f"Период: {start} — {end}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Тариф", "Кол-во запусков", "Выручка", "Цена тарифа"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for p in programs:
        ws.append([p["program_name"], p["launches_count"], p["total_revenue"], p["avg_price"]])

    # итоговая строка
    total_launches = sum(p["launches_count"] for p in programs)
    total_revenue = sum(p["total_revenue"] for p in programs)

    ws.append([])
    ws.append(["ИТОГО", total_launches, total_revenue, ""])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    for col in "BCD":
        ws.column_dimensions[col].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    machine_label = machine_id if machine_id else "all"
    filename = f"programs_report_{machine_label}_{start}_{end}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )